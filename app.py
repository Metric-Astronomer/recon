"""
Spaceship Reconciliation — Finance Team App
============================================
TO START: streamlit run app.py
"""

import os
import tempfile
import io
import csv
from datetime import datetime

import streamlit as st
import pandas as pd
import openpyxl

from reconcile import (
    load_commercial_rates,
    load_product_weights,
    load_invoice_from_csv,
    reconcile,
    generate_report,
)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR             = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RATES_FILE   = os.path.join(BASE_DIR, "Spaceship Shipping Fee Validation - Jan'26.xlsx")
DEFAULT_PRODUCT_FILE = os.path.join(BASE_DIR, "Spaceship Shipping Fee Validation - Jan'26.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE SETUP
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Spaceship Reconciliation",
    page_icon="📦",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# Minimal CSS — only touch things Streamlit can't do natively.
# No color rules here at all (Streamlit handles those reliably).
# We only hide branding and style the buttons.
st.markdown("""
<style>
    #MainMenu { visibility: hidden; }
    footer    { visibility: hidden; }
    header    { visibility: hidden; }

    .stDownloadButton > button {
        background-color: #1d4ed8 !important;
        color: #ffffff !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
        font-size: 1rem !important;
        width: 100% !important;
    }
    .stButton > button {
        background-color: #1d4ed8 !important;
        color: #ffffff !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
        font-size: 1rem !important;
        width: 100% !important;
    }
    .stButton > button:disabled {
        background-color: #94a3b8 !important;
        color: #ffffff !important;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# CACHING — rate table loaded once, lives in memory for the session
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def get_rates(rate_card_path: str):
    return load_commercial_rates(rate_card_path, "Commercial")

@st.cache_resource(show_spinner=False)
def get_product_weights():
    return load_product_weights(DEFAULT_PRODUCT_FILE, "Product LBH Master")


def normalize_uploaded_commercial_sheet(uploaded_file, tmpdir: str) -> str:
    """
    Convert an uploaded Commercial sheet (Excel or CSV) into a temporary
    workbook that always has one sheet named 'Commercial'.
    """
    filename = uploaded_file.name or "uploaded_rate_card"
    ext = os.path.splitext(filename)[1].lower()
    output_path = os.path.join(tmpdir, "uploaded_commercial_normalized.xlsx")

    if ext in (".xlsx", ".xlsm"):
        in_wb = openpyxl.load_workbook(
            io.BytesIO(uploaded_file.getvalue()),
            read_only=True,
            data_only=True,
        )
        source_ws = in_wb["Commercial"] if "Commercial" in in_wb.sheetnames else in_wb[in_wb.sheetnames[0]]

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Commercial"

        for r_idx, row in enumerate(source_ws.iter_rows(values_only=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                out_ws.cell(r_idx, c_idx, value)

        out_wb.save(output_path)
        in_wb.close()
        return output_path

    if ext == ".csv":
        text = uploaded_file.getvalue().decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Commercial"

        # Keep row 1 blank so the parser behavior matches the original workbook style.
        out_row = 2
        for row in reader:
            for c_idx, value in enumerate(row, start=1):
                out_ws.cell(out_row, c_idx, value)
            out_row += 1

        out_wb.save(output_path)
        return output_path

    raise ValueError("Step 3 supports only .xlsx, .xlsm, or .csv files for the Commercial sheet.")


# ─────────────────────────────────────────────────────────────────────────────
# HEADER — native Streamlit so text color is always correct
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("<div style='text-align:center; font-size:2.5rem; padding-top:1rem;'>📦</div>",
            unsafe_allow_html=True)
st.markdown("<h1 style='text-align:center; margin-top:0;'>Spaceship Reconciliation</h1>",
            unsafe_allow_html=True)
st.markdown("<p style='text-align:center; margin-bottom:1.5rem;'>Upload your monthly invoices → get the exceptions report instantly</p>",
            unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# PRELOAD RESOURCES
# ─────────────────────────────────────────────────────────────────────────────

with st.spinner("Loading rate card and product weights (first time only)..."):
    default_rates   = get_rates(DEFAULT_RATES_FILE)
    product_weights = get_product_weights()

st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — INVOICE FILES
# ─────────────────────────────────────────────────────────────────────────────

st.subheader("Step 1 — Upload Spaceship Invoice(s)")
st.caption("Download from the Spaceship portal. Upload all weekly files for the month together — they'll be combined automatically.")

invoice_files = st.file_uploader(
    label="Spaceship Invoice CSV(s)",
    type=["csv"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if invoice_files:
    total_rows = 0
    for f in invoice_files:
        try:
            preview_df = pd.read_csv(f)
            total_rows += len(preview_df)
            f.seek(0)
        except Exception:
            pass
    st.success(f"✓ {len(invoice_files)} file(s) uploaded — {total_rows:,} invoice rows")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — ORDER REPORT
# ─────────────────────────────────────────────────────────────────────────────

st.subheader("Step 2 — Upload Order/Mapping Report (optional but recommended)")
st.warning("Without this file, weight validation is skipped and you'll miss cases where Spaceship scanned the wrong weight.")

order_file = st.file_uploader(
    label="Order Report CSV",
    type=["csv"],
    accept_multiple_files=False,
    label_visibility="collapsed",
)

if order_file:
    try:
        order_preview = pd.read_csv(order_file)
        order_file.seek(0)
        st.success(f"✓ Mapping file uploaded — {len(order_preview):,} rows")
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — RATE CARD (OPTIONAL OVERRIDE)
# ─────────────────────────────────────────────────────────────────────────────

st.subheader("Step 3 — Upload Updated Rate Card (optional)")
st.caption("Upload only the updated Commercial sheet (Excel/CSV). If not uploaded, the default bundled rate card is used.")

rate_card_file = st.file_uploader(
    label="Updated Commercial Sheet (.xlsx/.xlsm/.csv)",
    type=["xlsx", "xlsm", "csv"],
    accept_multiple_files=False,
    label_visibility="collapsed",
)

if rate_card_file:
    st.success(f"✓ Updated rate card uploaded — {rate_card_file.name}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — MONTH
# ─────────────────────────────────────────────────────────────────────────────

st.subheader("Step 4 — Which month?")

col_month, col_year = st.columns(2)
with col_month:
    month_name = st.selectbox(
        "Month",
        ["January", "February", "March", "April", "May", "June",
         "July", "August", "September", "October", "November", "December"],
        index=datetime.now().month - 2 if datetime.now().month > 1 else 11,
    )
with col_year:
    year = st.selectbox("Year", [2025, 2026, 2027], index=1)

month_label = f"{month_name} {year}"

st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# RUN BUTTON
# ─────────────────────────────────────────────────────────────────────────────

can_run = bool(invoice_files)

run_clicked = st.button(
    f"▶  Run Reconciliation for {month_label}",
    disabled=not can_run,
    use_container_width=True,
)

if not can_run:
    st.caption("⬆ Upload at least one invoice file above to continue.")


# ─────────────────────────────────────────────────────────────────────────────
# RUN THE RECONCILIATION
# ─────────────────────────────────────────────────────────────────────────────

if run_clicked and invoice_files:

    progress_bar = st.progress(0, text="Loading invoice files...")

    with tempfile.TemporaryDirectory() as tmpdir:
        # Decide rate card for this run
        if rate_card_file:
            try:
                rate_card_path = normalize_uploaded_commercial_sheet(rate_card_file, tmpdir)
                # Cache key is file path string, so new uploads load separately.
                rates = get_rates(rate_card_path)
                selected_rate_card_name = f"{rate_card_file.name} (Commercial upload)"
            except Exception as e:
                st.error(f"Could not read Step 3 file. Upload only the Commercial sheet in Excel/CSV format. Error: {e}")
                st.stop()
        else:
            rates = default_rates
            selected_rate_card_name = os.path.basename(DEFAULT_RATES_FILE)

        # Save invoice CSVs
        csv_paths = []
        for f in invoice_files:
            path = os.path.join(tmpdir, f.name)
            with open(path, "wb") as out:
                out.write(f.read())
            csv_paths.append(path)

        # Save order/mapping file
        order_path = None
        if order_file:
            order_path = os.path.join(tmpdir, order_file.name)
            with open(order_path, "wb") as out:
                out.write(order_file.read())

        progress_bar.progress(25, text="Reading invoice data...")

        try:
            invoice_df = load_invoice_from_csv(csv_paths, order_path)
        except Exception as e:
            st.error(f"Failed to read invoice CSV: {e}")
            st.stop()

        progress_bar.progress(50, text="Running reconciliation...")

        try:
            result_df = reconcile(invoice_df, rates, product_weights)
        except Exception as e:
            st.error(f"Reconciliation failed: {e}")
            st.stop()

        progress_bar.progress(80, text="Generating Excel report...")

        output_path = os.path.join(tmpdir, f"reconciliation_{month_label.replace(' ', '_')}.xlsx")
        try:
            generate_report(result_df, output_path, month_label)
        except Exception as e:
            st.error(f"Failed to generate report: {e}")
            st.stop()

        progress_bar.progress(100, text="Done!")

        with open(output_path, "rb") as f:
            excel_bytes = f.read()

    progress_bar.empty()

    # ── RESULTS ──────────────────────────────────────────────────────────────

    st.divider()
    st.subheader(f"Results — {month_label}")
    st.caption(f"Rate card used: `{selected_rate_card_name}`")

    remark_counts  = result_df["Remarks"].value_counts().to_dict()
    matched        = remark_counts.get("Matched", 0)
    rate_diff      = remark_counts.get("Rate Diff", 0)
    weight_diff    = remark_counts.get("Weight Diff", 0)
    manual         = remark_counts.get("B2B - Manual Check", 0)
    no_rate        = remark_counts.get("No Rate", 0)

    actionable_df  = result_df[result_df["Remarks"].isin(["Rate Diff", "Weight Diff"])]
    total_variance = actionable_df["Charges Diff (HKD)"].sum()
    inr_rate       = 11.65
    variance_inr   = total_variance * inr_rate

    # Stat cards — only these use custom HTML (colored backgrounds need it)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(
            f"<div style='background:#e8f5e9;border-radius:8px;padding:1rem;text-align:center;'>"
            f"<div style='font-size:2rem;font-weight:900;color:#1b5e20;'>{matched:,}</div>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#2e7d32;'>Matched ✓</div></div>",
            unsafe_allow_html=True)
    with col2:
        st.markdown(
            f"<div style='background:#fff8e1;border-radius:8px;padding:1rem;text-align:center;'>"
            f"<div style='font-size:2rem;font-weight:900;color:#e65100;'>{rate_diff:,}</div>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#bf360c;'>Rate Diff ⚠</div></div>",
            unsafe_allow_html=True)
    with col3:
        st.markdown(
            f"<div style='background:#ffebee;border-radius:8px;padding:1rem;text-align:center;'>"
            f"<div style='font-size:2rem;font-weight:900;color:#b71c1c;'>{weight_diff:,}</div>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#c62828;'>Weight Diff 🔴</div></div>",
            unsafe_allow_html=True)
    with col4:
        st.markdown(
            f"<div style='background:#ede7f6;border-radius:8px;padding:1rem;text-align:center;'>"
            f"<div style='font-size:2rem;font-weight:900;color:#4a148c;'>{manual + no_rate:,}</div>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#6a1b9a;'>Manual Check 🔵</div></div>",
            unsafe_allow_html=True)

    st.write("")  # spacer

    if total_variance > 0:
        st.success(
            f"💰 **Total Recoverable Overcharge: HKD {total_variance:,.2f}**  \n"
            f"Approx **₹{variance_inr:,.0f}** at ₹{inr_rate}/HKD"
        )
    else:
        st.info("No overcharges detected. All charges match the rate card.")

    if manual > 0:
        st.warning(
            f"**{manual} B2B/bulk shipment(s)** need manual rate verification — "
            "large wholesale shipments not in the standard rate table. "
            "See the Exceptions sheet."
        )

    # Invoice breakdown table
    if "Invoice Number" in result_df.columns:
        st.markdown("**Breakdown by Invoice:**")
        inv_summary = (
            result_df[result_df["Remarks"].isin(["Rate Diff", "Weight Diff"])]
            .groupby("Invoice Number")
            .agg(
                Shipments=("Charges Diff (HKD)", "count"),
                Variance_HKD=("Charges Diff (HKD)", "sum"),
            )
            .reset_index()
        )
        inv_summary["Variance_HKD"] = inv_summary["Variance_HKD"].round(2)
        inv_summary.columns = ["Invoice", "Exception Rows", "Variance (HKD)"]
        st.dataframe(inv_summary, use_container_width=True, hide_index=True)

    # Download button
    st.divider()
    filename = f"reconciliation_{month_label.replace(' ', '_')}.xlsx"
    st.download_button(
        label=f"⬇  Download Full Report — {filename}",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Report has 3 sheets: Summary · Exceptions (send to Spaceship) · Full Detail")


# ─────────────────────────────────────────────────────────────────────────────
# FOOTER — how to update rate card
# ─────────────────────────────────────────────────────────────────────────────

with st.expander("ℹ️  How to update the rate card or add new products"):
    st.markdown("""
**Rate card changed?** (Spaceship sends a new commercial rate sheet)
1. Upload only the updated **Commercial** sheet in **Step 3** (`.xlsx`, `.xlsm`, or `.csv`)
2. Run reconciliation — no restart needed

**New SKU launched?**
1. Add the SKU + actual weight (kg) to the `Product LBH Master` sheet in the master Excel
2. Restart the app

**What files do you need each month?**
- **Spaceship invoices** — Download from Spaceship portal → Reports → Billing → export as CSV.
  If the portal only allows 7 days at a time, download 4–5 weekly files and upload them all together.
- **Order/Mapping report** — Download from Noise/Shopify → Orders → Export.
  Needs columns: `Product SKU Code` and `Qty` (one row per shipment, in the same order as the invoice).
    """)
