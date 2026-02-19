"""
Spaceship Shipping Fee Reconciliation Engine
=============================================

What this script does (like you're 12 years old):
--------------------------------------------------
Imagine you ordered something online and the delivery company charges you $50.
But YOUR rate card (the deal you negotiated with them) says it should only be $40.
That $10 difference is money you can get back — IF you catch it.

This script is a robot that:
1. Reads the Spaceship invoice (what they charged you)
2. Reads YOUR rate card (Commercial table — what they SHOULD charge)
3. Reads the product weight list (so we know what each item actually weighs)
4. For every single shipment, asks: "Did Spaceship charge the right amount?"
5. Flags every mismatch with the exact dollar difference
6. Spits out a clean Excel report grouped by invoice, carrier, and issue type

HOW THE LOOKUP WORKS (the key magic):
--------------------------------------
Every row in the invoice has: Carrier + From + To + Weight
Every row in the rate table has: Carrier + From + To + Weight → Expected Price

To find the expected price, we do a "ceiling match":
- If the package weighs 0.34 kg, the nearest rate tier above it is 0.5 kg
- So we charge at the 0.5 kg rate
- Think of it like a taxi meter: 0-0.5 km = $5, 0.5-1 km = $8... you pay the band you fall into

TWO TYPES OF VARIANCE:
-----------------------
1. RATE DIFF   → Same weight, but Spaceship used an old/wrong rate from their side
2. WEIGHT DIFF → Spaceship said the package weighs 11 kg when it actually weighs 0.5 kg (!!!)
                 This is a data entry / scanning error at their end

INPUTS (what you need each month):
------------------------------------
  - Spaceship invoice CSV(s)  : Download from Spaceship portal (can be multiple weekly files)
  - Order Report CSV          : Download from Noise/Shopify portal — gives SKU + Qty per order
  - commercial_rates.xlsx     : The negotiated rate card (update when rates change)
  - product_master.xlsx       : SKU → actual weight (update when new products launch)

OUTPUT:
-------
  A single Excel file: reconciliation_YYYY_MM.xlsx with:
  - Sheet "Summary"     : Total variance by invoice + by carrier
  - Sheet "Exceptions"  : Only the rows with differences (actionable list to send to Spaceship)
  - Sheet "Full Detail" : Every single row (audit trail)

USAGE:
------
  # Option A: Run against the master Excel directly (proof of concept / re-run)
  python3 reconcile.py --source-excel "Spaceship Shipping Fee Validation - Jan'26.xlsx"

  # Option B: Run against fresh CSV exports each month (normal monthly run)
  python3 reconcile.py \\
      --invoice-csv  inputs/invoice_week1.csv inputs/invoice_week2.csv \\
      --order-csv    inputs/order_report.csv \\
      --rates-excel  inputs/commercial_rates.xlsx \\
      --product-excel inputs/product_master.xlsx \\
      --month        2026-02
"""

import argparse
import math
import os
import sys
from datetime import datetime
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: LOAD THE COMMERCIAL RATE TABLE
# ─────────────────────────────────────────────────────────────────────────────

def load_commercial_rates(excel_path: str, sheet_name: str = "Commercial") -> dict:
    """
    Load the 170k-row rate table into a fast lookup dict.

    Think of it like building a phone book:
    Instead of searching all 170k rows every time, we pre-sort by
    (Carrier, To) → sorted list of (weight_threshold, freight, fuel, other)

    So a lookup is O(log n) binary search, not O(n) full scan.
    That's the difference between 1ms and 170ms per row.

    HOW THE ORIGINAL EXCEL FORMULA WORKS:
    =======================================
    The key formula in the sheet is:
      =MIN(SP_Freight, INDEX(Commercial!I:I, MATCH(1,
           (Commercial!A:A = Courier) *
           (Commercial!F:F = Noise_Weight_KG) *
           (Commercial!E:E = To_Country), 0)))

    This means the lookup key is ONLY: Carrier + To + Weight.
    The "From" country is NOT part of the lookup! (All shipments are from HK anyway.)

    Returns:
        dict: {(carrier, to_country): [(weight, freight, fuel, other_surcharges, total_rate), ...]}
              Each inner list is sorted by weight ascending (for binary-search ceiling match).
    """
    print("Loading commercial rate table... (this is the big 170k-row file, give it a moment)")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rates = defaultdict(list)
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: LookupValue(A), Carrier(B), Service(C), From(D), To(E), Weight(F),
        #          TaxTerm(G), Rate(HKD)(H), ShippingFee(I), FuelSurcharge(J), OtherSurcharges(K) ...
        lookup_val = row[0]   # A: carrier display name used in invoice
        to_c       = row[4]   # E: To country
        weight     = row[5]   # F: Weight tier (kg)
        rate_total = row[7]   # H: Rate (HKD) — all-in
        freight    = row[8]   # I: ShippingFee (base freight)
        fuel       = row[9]   # J: FuelSurcharge
        other      = row[10]  # K: OtherSurcharges

        # Skip header row or invalid rows
        if not lookup_val or lookup_val in ("Lookup Value", "#N/A") or not isinstance(weight, (int, float)):
            skipped += 1
            continue

        # KEY: Only (carrier, to_country) — no From in the lookup
        key = (str(lookup_val).strip(), str(to_c).strip())
        rates[key].append((
            float(weight),
            float(freight)    if freight    else 0.0,
            float(fuel)       if fuel       else 0.0,
            float(other)      if other      else 0.0,
            float(rate_total) if rate_total else 0.0,
        ))

    # Sort each list by weight so we can binary-search for ceiling matches
    for key in rates:
        rates[key].sort(key=lambda x: x[0])

    wb.close()

    total_keys = len(rates)
    total_rows = sum(len(v) for v in rates.values())
    print(f"  ✓ Loaded {total_rows:,} rate rows across {total_keys:,} carrier/destination combinations")
    print(f"  ✓ Skipped {skipped:,} header/invalid rows")

    return dict(rates)


def _ceiling_half(value: float) -> float:
    """
    Round UP to the nearest 0.5 kg — exactly what CEILING.MATH(x, 0.5) does in Excel.

    Examples:
      0.27 → 0.5
      0.35 → 0.5
      0.50 → 0.5
      0.51 → 1.0
      1.00 → 1.0
      1.01 → 1.5
      11.0 → 11.0
    """
    return math.ceil(value / 0.5) * 0.5


def lookup_rate(rates: dict, carrier: str, to_c: str, noise_weight_kg: float):
    """
    Find the expected rate for a given carrier + to_country + weight.

    Since noise_weight_kg is already CEILING'd to 0.5 steps (e.g. 0.5, 1.0, 1.5...),
    we look for the first tier in the rate table that is >= noise_weight_kg.

    Think of it like a staircase: the rate table has steps at 0.5, 1.0, 3.5, 4.5...
    If you need the rate for 1.0 kg, you pick the first step at or above 1.0.

    Returns:
        (freight, fuel, other, total_rate, matched_weight_tier) or None if not found
    """
    key = (carrier, to_c)
    tiers = rates.get(key)
    if not tiers:
        return None

    # Binary search for ceiling: find the LOWEST tier that is >= noise_weight_kg
    lo, hi = 0, len(tiers) - 1
    result = None
    while lo <= hi:
        mid = (lo + hi) // 2
        tier_weight = tiers[mid][0]
        if tier_weight >= noise_weight_kg - 1e-9:
            result = tiers[mid]
            hi = mid - 1  # keep searching left for a lower qualifying tier
        else:
            lo = mid + 1

    if result is None:
        # Weight exceeds max tier in the table — use the highest available
        result = tiers[-1]

    return result[1], result[2], result[3], result[4], result[0]


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: LOAD THE PRODUCT WEIGHT MASTER
# ─────────────────────────────────────────────────────────────────────────────

def load_product_weights(excel_path: str, sheet_name: str = "Product LBH Master") -> dict:
    """
    Load SKU → actual weight (kg) mapping.

    The product master has:
    - SKU code (e.g. "wrb-rg-luna-7-sgld-2")
    - Actual-Weight in kg (e.g. 0.35)

    Returns:
        dict: {sku_lowercase: weight_kg}
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    weights = {}
    header_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[1] == 'SKU':
            header_found = True
            continue
        if not header_found:
            continue
        sku = row[1]
        actual_weight = row[6]  # Column G: Actual - Weight
        if sku and actual_weight:
            weights[str(sku).strip().lower()] = float(actual_weight)

    wb.close()
    print(f"  ✓ Loaded {len(weights):,} product SKU weights")
    return weights


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: LOAD INVOICE DATA
# ─────────────────────────────────────────────────────────────────────────────

# Column positions in "Shipping Freight - Cost" sheet (0-indexed)
# These match the exact headers in the Spaceship invoice sheet
INVOICE_COLS = {
    "tracking_number":      0,
    "date":                 1,
    "courier":              2,
    "cust_name":            3,
    "trn_id":               4,
    "order_ref":            5,
    "preview_weight":       6,
    "final_weight":         7,   # What Spaceship claims the weight is
    "piece":                8,
    "from_country":         9,
    "to_country":           10,
    "sp_freight":           11,  # Spaceship billed: freight
    "sp_fuel":              12,  # Spaceship billed: fuel
    "sp_seasonal":          13,
    "sp_residential":       14,
    "sp_remote":            15,
    "sp_tariff":            16,
    "sp_addr_correction":   17,
    "sp_dt_handling":       18,
    "sp_addl_handling":     19,
    "sp_us_inbound":        20,
    "sp_total":             21,  # Total charged by Spaceship
    "invoice_number":       22,
    "invoice_date":         23,
    "country_name":         24,
    "duplicate_awb":        25,
    "sku":                  26,  # Added from order report
    "qty":                  27,  # Added from order report
    "noise_weight_grams":   28,  # Expected weight in grams (qty × product weight × 1000)
    "noise_weight_kg":      29,  # Expected weight rounded to nearest rate tier
    "noise_freight":        30,
    "noise_fuel":           31,
    "noise_seasonal":       32,
    "noise_residential":    33,
    "noise_remote":         34,
    "noise_tariff":         35,
    "noise_addr_correction":36,
    "noise_dt_handling":    37,
    "noise_addl_handling":  38,
    "noise_elec_label":     39,
    "noise_rounding":       40,
    "noise_total":          41,
    "charges_diff":         42,
    "remarks":              43,
    "addon_comments":       44,
    "weight_diff":          45,
    "freight_diff":         46,
    "fuel_diff":            47,
    "fsc_pct":              48,
    "b2b_invoice_ref":      49,
}


def load_invoice_from_excel(excel_path: str, sheet_name: str = "Shipping Freight - Cost") -> pd.DataFrame:
    """
    Load the pre-filled invoice sheet from the master Excel.
    This is the "proof of concept" mode — we read the already-reconciled data
    and RE-COMPUTE everything to verify our logic matches.
    """
    print(f"Loading invoice data from Excel sheet '{sheet_name}'...")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if row[0] == "Tracking number":
            header_row_idx = i
            continue
        if header_row_idx is None:
            continue
        # Skip empty rows
        if not row[0]:
            continue
        rows.append(row)

    wb.close()

    col_names = list(INVOICE_COLS.keys())
    # Only take as many columns as we defined
    max_col = max(INVOICE_COLS.values()) + 1
    clean_rows = []
    for r in rows:
        padded = list(r) + [None] * max_col
        clean_rows.append(padded[:max_col])

    df = pd.DataFrame(clean_rows, columns=[
        list(INVOICE_COLS.keys())[i] if i < len(INVOICE_COLS) else f"col_{i}"
        for i in range(max_col)
    ])

    print(f"  ✓ Loaded {len(df):,} invoice rows")
    return df


def _clean_csv(path: str) -> pd.DataFrame:
    """
    Read a CSV and fix the two most common real-world messiness issues
    from the Spaceship portal export:

    1. Extra spaces in column names: " Final weight " → "Final weight"
    2. Dash placeholders for zero: "  -   " → 0.0
       (Spaceship uses dashes instead of 0 for empty charge columns)
    """
    df = pd.read_csv(path, dtype=str)

    # Strip leading/trailing whitespace from ALL column names
    df.columns = [c.strip() for c in df.columns]

    # Strip whitespace from every cell value, then replace "-" variants with "0"
    for col in df.columns:
        df[col] = df[col].str.strip()
        # "-", "- ", " -", "  -   " are all zero
        df[col] = df[col].replace(r'^\s*-\s*$', '0', regex=True)

    return df


def load_invoice_from_csv(csv_paths: list, order_csv_path: str = None) -> pd.DataFrame:
    """
    Load Spaceship invoice from one or more CSV exports (weekly files → merge into monthly).

    SPACESHIP EXPORT COLUMNS (standard Spaceship portal CSV format):
      Tracking number, Date, Courier, Cust. Name, Trn. ID, Order ref.,
      Preview weight, Final weight, Piece, From, To,
      Freight charges, Fuel surcharges, Seasonal surcharge,
      Residential surcharges, Remote Area surcharges, Tariff,
      Address correction, D&T Handling fee, Additional Handling,
      US Inbound Processing Fee, Total Cost (or "Total Cost - Spaceship"),
      Invoice Number, Invoice Date

    ORDER REPORT / MAPPING COLUMNS (from Noise/Shopify portal):
      Option A — with Order ref (preferred):
        Order ref., Product SKU Code, Qty
      Option B — positional (same row order as invoice, no Order ref column):
        Product SKU Code, Qty
        (rows are joined 1-to-1 by position)

    We join on Order ref. when available, otherwise join by row position.
    """
    print(f"Loading {len(csv_paths)} invoice CSV file(s)...")

    dfs = []
    for path in csv_paths:
        df = _clean_csv(path)
        dfs.append(df)
        print(f"  ✓ Loaded {len(df):,} rows from {os.path.basename(path)}")

    invoice_df = pd.concat(dfs, ignore_index=True)
    print(f"  ✓ Combined total: {len(invoice_df):,} invoice rows")

    # Rename columns to our internal names.
    # We handle both "Total Cost" and "Total Cost - Spaceship" (portal uses the longer name).
    col_map = {
        "Tracking number":           "tracking_number",
        "Date":                      "date",
        "Courier":                   "courier",
        "Cust. Name":                "cust_name",
        "Trn. ID":                   "trn_id",
        "Order ref.":                "order_ref",
        "Preview weight":            "preview_weight",
        "Final weight":              "final_weight",
        "Piece":                     "piece",
        "From":                      "from_country",
        "To":                        "to_country",
        "Freight charges":           "sp_freight",
        "Fuel surcharges":           "sp_fuel",
        "Seasonal surcharge":        "sp_seasonal",
        "Residential surcharges":    "sp_residential",
        "Remote Area surcharges":    "sp_remote",
        "Tariff":                    "sp_tariff",
        "Address correction":        "sp_addr_correction",
        "D&T Handling fee":          "sp_dt_handling",
        "Additional Handling":       "sp_addl_handling",
        "US Inbound Processing Fee": "sp_us_inbound",
        "Total Cost":                "sp_total",
        "Total Cost - Spaceship":    "sp_total",   # alternate name from portal export
        "Invoice Number":            "invoice_number",
        "Invoice Date":              "invoice_date",
        "Country Name":              "country_name",
    }
    invoice_df = invoice_df.rename(columns=col_map)

    # Convert all numeric charge columns from string to float
    numeric_cols = [
        "preview_weight", "final_weight", "piece",
        "sp_freight", "sp_fuel", "sp_seasonal", "sp_residential", "sp_remote",
        "sp_tariff", "sp_addr_correction", "sp_dt_handling",
        "sp_addl_handling", "sp_us_inbound", "sp_total",
    ]
    for col in numeric_cols:
        if col in invoice_df.columns:
            invoice_df[col] = pd.to_numeric(invoice_df[col], errors="coerce").fillna(0.0)

    # Join with order report / mapping file if provided
    if order_csv_path:
        print(f"Joining with order/mapping file from {os.path.basename(order_csv_path)}...")
        order_df = _clean_csv(order_csv_path)

        # Normalise column names (strip + lowercase for matching)
        order_df.columns = [c.strip() for c in order_df.columns]

        # Build a flexible column map — handle any reasonable variation
        order_col_map = {}
        for col in order_df.columns:
            cl = col.lower().strip()
            if cl in ("order ref.", "order ref", "order_ref", "order reference"):
                order_col_map[col] = "order_ref"
            elif cl in ("product sku code", "sku code", "sku", "product sku", "product_sku_code"):
                order_col_map[col] = "sku"
            elif cl in ("qty", "quantity", "units"):
                order_col_map[col] = "qty"
        order_df = order_df.rename(columns=order_col_map)

        has_order_ref = "order_ref" in order_df.columns
        has_sku       = "sku"       in order_df.columns
        has_qty       = "qty"       in order_df.columns

        if has_sku and has_qty:
            if has_order_ref:
                # Option A: proper join on Order ref.
                order_df["order_ref"] = order_df["order_ref"].str.strip()
                order_df = order_df[["order_ref", "sku", "qty"]].drop_duplicates("order_ref")
                invoice_df["order_ref"] = invoice_df["order_ref"].astype(str).str.strip()
                invoice_df = invoice_df.merge(order_df, on="order_ref", how="left")
                print(f"  ✓ Joined on Order ref. — {order_df.shape[0]:,} mappings")
            else:
                # Option B: positional join (row N of mapping = row N of invoice)
                order_df = order_df[["sku", "qty"]].reset_index(drop=True)
                invoice_df = invoice_df.reset_index(drop=True)
                invoice_df["sku"] = order_df["sku"]
                invoice_df["qty"] = order_df["qty"]
                print(f"  ✓ Joined positionally (no Order ref. column found) — {len(order_df):,} rows matched")
        else:
            missing = [c for c, f in [("sku", has_sku), ("qty", has_qty)] if not f]
            print(f"  ⚠ Order/mapping file missing columns: {missing}. Skipping SKU join.")
            invoice_df["sku"] = None
            invoice_df["qty"] = None
    else:
        invoice_df["sku"] = None
        invoice_df["qty"] = None

    return invoice_df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: THE RECONCILIATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def reconcile(df: pd.DataFrame, rates: dict, product_weights: dict) -> pd.DataFrame:
    """
    The heart of the automation. Replicates the exact Excel logic.

    EXACT EXCEL LOGIC (reverse-engineered from the formulas):
    ===========================================================

    STEP 1 — Compute Noise Weight (grams):
      = MIN(SP_Final_Weight_kg, Product_Actual_Weight_kg × Qty) × 1000
      Why MIN? If Spaceship says a package is 0.27 kg but the product is 0.35 kg,
      we give them benefit of doubt and use 0.27 (maybe they measured differently).
      But if they say 11 kg and the product is 0.35 kg — we use 0.35 (clear error).

    STEP 2 — Compute Noise Weight KG (the rate tier to look up):
      = CEILING(Noise_Weight_grams / 1000, 0.5)
      This rounds UP to the nearest 0.5 kg step.
      Example: 270g = 0.27 kg → rounds to 0.5. The 0.5 kg rate band applies.

    STEP 3 — Look up expected rate in Commercial table:
      MATCH key: Carrier Name (A column) + Destination Country (E column) + Weight (F column)
      NOTE: Origin/From country is NOT in the lookup — it's implied to always be HK.

    STEP 4 — Expected charge = MIN(SP_charge, rate_table_charge):
      = MIN(SP_Freight, Looked_Up_Freight)
      Why MIN? We only dispute OVERCHARGES. If Spaceship charges LESS than the
      rate card (a discount), we accept it and show it as Matched.

    STEP 5 — Classify:
      Weight_Diff_check = CEILING(Noise_Weight_KG, 0.5) - CEILING(SP_Final_Weight, 0.5)
      ≠ 0  → "Weight Diff"  (e.g. product is 0.35 kg but Spaceship scanned 11 kg)
      = 0 and charges_diff ≈ 0 → "Matched"
      = 0 and charges_diff ≠ 0 → "Rate Diff"

    Classifications:
      - "Matched"      → Everything correct ✓
      - "Rate Diff"    → Same weight tier, wrong price (Spaceship used old/wrong rate)
      - "Weight Diff"  → Spaceship charged for wrong weight (much more than product weighs)
      - "No Rate"      → Carrier/destination not in our rate table (needs manual check)
      - "Duplicate"    → Same tracking number billed twice
    """
    print(f"\nRunning reconciliation on {len(df):,} rows...")

    results = []

    # Track duplicates (same tracking number appearing twice)
    tracking_counts = df["tracking_number"].value_counts()
    duplicates = set(tracking_counts[tracking_counts > 1].index)

    matched = rate_diff = weight_diff = no_rate = dupe = 0
    TOLERANCE = 0.02  # HKD — rounding tolerance (< 2 cents is not a real diff)

    for idx, row in df.iterrows():
        tracking = str(row.get("tracking_number", "")).strip()
        courier  = str(row.get("courier", "")).strip()
        from_c   = str(row.get("from_country", "")).strip()
        to_c     = str(row.get("to_country", "")).strip()
        sku      = str(row.get("sku", "")).strip().lower() if row.get("sku") else ""
        qty      = _safe_float(row.get("qty")) or 1.0

        # Spaceship's billed amounts
        sp_final_weight  = _safe_float(row.get("final_weight"))
        sp_freight       = _safe_float(row.get("sp_freight"))
        sp_fuel          = _safe_float(row.get("sp_fuel"))
        sp_seasonal      = _safe_float(row.get("sp_seasonal"))
        sp_residential   = _safe_float(row.get("sp_residential"))
        sp_remote        = _safe_float(row.get("sp_remote"))
        sp_tariff        = _safe_float(row.get("sp_tariff"))
        sp_addr_corr     = _safe_float(row.get("sp_addr_correction"))
        sp_dt_handling   = _safe_float(row.get("sp_dt_handling"))
        sp_addl_handling = _safe_float(row.get("sp_addl_handling"))
        sp_us_inbound    = _safe_float(row.get("sp_us_inbound"))
        sp_total         = _safe_float(row.get("sp_total"))

        # ── STEP 1: Noise Weight (grams) ────────────────────────────────────
        # = MIN(SP_Final_Weight, product_weight × qty) × 1000
        product_weight_kg = product_weights.get(sku) if sku else None
        if product_weight_kg and sp_final_weight > 0:
            noise_weight_raw = min(sp_final_weight, product_weight_kg * qty)
        elif sp_final_weight > 0:
            noise_weight_raw = sp_final_weight
        else:
            noise_weight_raw = 0.0

        noise_weight_grams = round(noise_weight_raw * 1000, 2)

        # ── STEP 2: Noise Weight KG — CEILING to 0.5 ────────────────────────
        # = CEILING(noise_weight_grams / 1000, 0.5)
        noise_weight_kg = _ceiling_half(noise_weight_grams / 1000) if noise_weight_grams > 0 else 0.0

        # ── STEP 3: Rate table lookup (key: Carrier + To + Noise_Weight_KG) ─
        lookup_result = lookup_rate(rates, courier, to_c, noise_weight_kg) if noise_weight_kg > 0 else None

        # ── STEP 4 & 5: Compare and classify ────────────────────────────────
        # Flag B2B / bulk shipments: large weight + SKU not in product master
        is_bulk_b2b = (not product_weight_kg) and (sp_final_weight > 30)

        if lookup_result is None:
            remarks = "No Rate"
            exp_freight = exp_fuel = noise_total = None
            charges_diff = None
            weight_diff_val = None
            no_rate += 1
        elif is_bulk_b2b:
            # B2B bulk shipment: can't auto-validate, need manual rate check
            remarks = "B2B - Manual Check"
            rate_freight, rate_fuel, rate_other, rate_total, matched_tier = lookup_result
            exp_freight = min(sp_freight, rate_freight)
            exp_fuel    = min(sp_fuel, rate_fuel)
            noise_total = round(exp_freight + exp_fuel + sp_seasonal + sp_residential + sp_remote + sp_tariff + sp_addr_corr + sp_dt_handling + sp_addl_handling + sp_us_inbound, 4)
            charges_diff = round(sp_total - noise_total, 4)
            sp_weight_ceil = _ceiling_half(sp_final_weight) if sp_final_weight > 0 else 0.0
            weight_diff_val = round(noise_weight_kg - sp_weight_ceil, 4)
            no_rate += 1
        else:
            rate_freight, rate_fuel, rate_other, rate_total, matched_tier = lookup_result

            # MIN rule: expected = min(what Spaceship charged, what rate table says)
            # → only flag overcharges, accept undercharges (discounts) silently
            exp_freight = min(sp_freight, rate_freight)
            exp_fuel    = min(sp_fuel, rate_fuel)

            # Expected total: use our expected freight+fuel, keep all other charges as-is
            # (pass-through surcharges: seasonal, residential, remote, tariff, etc.)
            noise_total = round(
                exp_freight + exp_fuel
                + sp_seasonal + sp_residential + sp_remote
                + sp_tariff + sp_addr_corr + sp_dt_handling
                + sp_addl_handling + sp_us_inbound,
                4
            )

            charges_diff = round(sp_total - noise_total, 4)

            # Weight diff check:
            # AT = CEILING(Noise_Weight_KG, 0.5) - CEILING(SP_Final_Weight, 0.5)
            # Since noise_weight_kg is already CEILING'd to 0.5, it equals CEILING(noise_raw, 0.5)
            sp_weight_ceil = _ceiling_half(sp_final_weight) if sp_final_weight > 0 else 0.0
            weight_diff_val = round(noise_weight_kg - sp_weight_ceil, 4)

            # Classify — ORDER MATTERS: charges first, then weight, then rate
            is_dup = tracking in duplicates

            if is_dup:
                remarks = "Duplicate"
                dupe += 1
            elif abs(charges_diff) <= TOLERANCE:
                # Charges match → Matched, regardless of weight tier difference
                # (Spaceship may state a different weight but we still got the same price)
                remarks = "Matched"
                matched += 1
            elif abs(weight_diff_val) > 0.01:
                # Charges differ AND weight tier is wrong → Weight Diff
                # This means Spaceship scanned a much higher weight than the product is
                remarks = "Weight Diff"
                weight_diff += 1
            else:
                # Charges differ but weight tier is correct → Rate Diff
                # Same weight, wrong price — Spaceship used an old/incorrect rate
                remarks = "Rate Diff"
                rate_diff += 1

        # ── Build result row ─────────────────────────────────────────────────
        result = {
            # === SPACESHIP BILLED DATA ===
            "Tracking Number":        tracking,
            "Date":                   row.get("date"),
            "Courier":                courier,
            "Order Ref":              row.get("order_ref"),
            "Invoice Number":         row.get("invoice_number"),
            "Invoice Date":           row.get("invoice_date"),
            "From":                   from_c,
            "To":                     to_c,
            "Country Name":           row.get("country_name"),
            "SKU":                    row.get("sku"),
            "Qty":                    int(qty),
            "SP Final Weight (kg)":   sp_final_weight,
            "SP Freight":             sp_freight,
            "SP Fuel Surcharge":      sp_fuel,
            "SP Seasonal":            sp_seasonal,
            "SP Residential":         sp_residential,
            "SP Remote Area":         sp_remote,
            "SP Total":               sp_total,
            # === EXPECTED (NOISE) CHARGES ===
            "Product Weight (kg)":    product_weight_kg,
            "Noise Weight (g)":       noise_weight_grams,
            "Noise Weight Tier (kg)": noise_weight_kg,
            "Expected Freight":       exp_freight,
            "Expected Fuel":          exp_fuel,
            "Expected Total":         noise_total,
            # === VARIANCE ===
            "Charges Diff (HKD)":     charges_diff,
            "Weight Diff (kg)":       weight_diff_val,
            "Freight Diff":           round(sp_freight - exp_freight, 4) if exp_freight is not None else None,
            "Fuel Diff":              round(sp_fuel - exp_fuel, 4) if exp_fuel is not None else None,
            "Remarks":                remarks,
        }
        results.append(result)

    result_df = pd.DataFrame(results)

    print(f"\n  ─── RECONCILIATION RESULTS ───")
    print(f"  Matched      : {matched:,} rows  (everything correct)")
    print(f"  Rate Diff    : {rate_diff:,} rows  (wrong price, same weight)")
    print(f"  Weight Diff  : {weight_diff:,} rows  (wrong weight charged)")
    print(f"  No Rate      : {no_rate:,} rows  (route not in rate table)")
    print(f"  Duplicate    : {dupe:,} rows  (same tracking appears twice)")
    print(f"  Total        : {len(result_df):,} rows")

    total_variance = result_df["Charges Diff (HKD)"].dropna().sum()
    print(f"\n  Total Variance (overcharge): HKD {total_variance:,.2f}")
    inr_rate = 11.65
    print(f"  In INR (~₹{inr_rate}/HKD): ₹{total_variance * inr_rate:,.0f}")
    print(f"  (Negative = Spaceship overcharged; Positive = they undercharged)")

    return result_df


def _safe_float(val) -> float:
    """Convert a value to float safely, returning 0.0 if None/blank."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: GENERATE EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────

# Color palette
COLORS = {
    "header_bg":    "1F4E79",   # Deep blue
    "header_font":  "FFFFFF",   # White
    "matched":      "E2EFDA",   # Light green
    "rate_diff":    "FFF2CC",   # Light yellow
    "weight_diff":  "FCE4D6",   # Light orange
    "no_rate":      "F4CCCC",   # Light red
    "duplicate":    "EAD1DC",   # Light pink
    "section_bg":   "D6E4F0",   # Light blue (section headers)
    "total_bg":     "BDD7EE",   # Medium blue (totals)
    "grand_total":  "1F4E79",   # Deep blue (grand total)
}

REMARK_COLORS = {
    "Matched":     COLORS["matched"],
    "Rate Diff":   COLORS["rate_diff"],
    "Weight Diff": COLORS["weight_diff"],
    "No Rate":     COLORS["no_rate"],
    "Duplicate":   COLORS["duplicate"],
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _bold_font(color: str = "000000", size: int = 10) -> Font:
    return Font(bold=True, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_report(result_df: pd.DataFrame, output_path: str, month_label: str = ""):
    """
    Write a clean, color-coded Excel report with 3 sheets:
    1. Summary    — totals by invoice and by carrier
    2. Exceptions — only rows with differences
    3. Full Detail — every row (complete audit trail)
    """
    print(f"\nGenerating Excel report: {output_path}")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    _write_summary_sheet(wb, result_df, month_label)

    # ── Sheet 2: Exceptions ──────────────────────────────────────────────────
    exceptions_df = result_df[result_df["Remarks"] != "Matched"].copy()
    _write_detail_sheet(wb, exceptions_df, sheet_title="Exceptions",
                        description="Rows where Spaceship's charges differ from expected. Send this list to Spaceship for credit/debit notes.")

    # ── Sheet 3: Full Detail ─────────────────────────────────────────────────
    _write_detail_sheet(wb, result_df, sheet_title="Full Detail",
                        description="Complete audit trail — every invoice row with expected vs actual charges.")

    # Remove default blank sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"  ✓ Report saved to: {output_path}")
    print(f"  ✓ Sheets: Summary | Exceptions ({len(exceptions_df):,} rows) | Full Detail ({len(result_df):,} rows)")


def _write_summary_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, month_label: str):
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws["A1"] = f"Spaceship Shipping Fee Reconciliation — {month_label}"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["header_bg"])
    ws.merge_cells("A1:N1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")

    current_row = 4

    # ── Section A: Summary by Invoice ─────────────────────────────────────
    ws.cell(current_row, 1, "SUMMARY BY INVOICE").font = _bold_font(COLORS["header_font"], 11)
    ws.cell(current_row, 1).fill = _fill(COLORS["header_bg"])
    ws.merge_cells(f"A{current_row}:H{current_row}")
    current_row += 1

    inv_headers = ["Invoice Number", "Remarks", "Shipments", "SP Total (HKD)", "Expected Total (HKD)", "Variance (HKD)", "Variance %"]
    for col, h in enumerate(inv_headers, 1):
        cell = ws.cell(current_row, col, h)
        cell.fill = _fill(COLORS["section_bg"])
        cell.font = _bold_font(size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()
    current_row += 1

    for inv_num, inv_group in df.groupby("Invoice Number"):
        for remark, remark_group in inv_group.groupby("Remarks"):
            sp_total    = remark_group["SP Total"].sum()
            noise_total = remark_group["Expected Total"].sum()
            variance    = remark_group["Charges Diff (HKD)"].sum()
            variance_pct = (variance / sp_total * 100) if sp_total else 0
            row_data = [inv_num, remark, len(remark_group), round(sp_total,2), round(noise_total,2), round(variance,2), f"{variance_pct:.2f}%"]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(current_row, col, val)
                cell.fill = _fill(REMARK_COLORS.get(remark, "FFFFFF"))
                cell.border = _border()
                cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
            current_row += 1

        # Invoice subtotal
        inv_sp_total     = inv_group["SP Total"].sum()
        inv_noise_total  = inv_group["Expected Total"].sum()
        inv_variance     = inv_group["Charges Diff (HKD)"].sum()
        inv_pct          = (inv_variance / inv_sp_total * 100) if inv_sp_total else 0
        subtotal_data = [f"{inv_num} TOTAL", "", len(inv_group), round(inv_sp_total,2), round(inv_noise_total,2), round(inv_variance,2), f"{inv_pct:.2f}%"]
        for col, val in enumerate(subtotal_data, 1):
            cell = ws.cell(current_row, col, val)
            cell.fill = _fill(COLORS["total_bg"])
            cell.font = _bold_font(size=9)
            cell.border = _border()
        current_row += 1

    # Grand total
    grand_sp    = df["SP Total"].sum()
    grand_noise = df["Expected Total"].sum()
    grand_var   = df["Charges Diff (HKD)"].sum()
    grand_pct   = (grand_var / grand_sp * 100) if grand_sp else 0
    grand_data  = ["GRAND TOTAL", "", len(df), round(grand_sp,2), round(grand_noise,2), round(grand_var,2), f"{grand_pct:.2f}%"]
    for col, val in enumerate(grand_data, 1):
        cell = ws.cell(current_row, col, val)
        cell.fill = _fill(COLORS["grand_total"])
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.border = _border()
    current_row += 2

    # ── Section B: Summary by Carrier ─────────────────────────────────────
    ws.cell(current_row, 1, "SUMMARY BY CARRIER").font = _bold_font(COLORS["header_font"], 11)
    ws.cell(current_row, 1).fill = _fill(COLORS["header_bg"])
    ws.merge_cells(f"A{current_row}:H{current_row}")
    current_row += 1

    for col, h in enumerate(inv_headers, 1):
        cell = ws.cell(current_row, col, h.replace("Invoice Number", "Carrier"))
        cell.fill = _fill(COLORS["section_bg"])
        cell.font = _bold_font(size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()
    current_row += 1

    for carrier, c_group in df.groupby("Courier"):
        for remark, remark_group in c_group.groupby("Remarks"):
            sp_total    = remark_group["SP Total"].sum()
            noise_total = remark_group["Expected Total"].sum()
            variance    = remark_group["Charges Diff (HKD)"].sum()
            variance_pct = (variance / sp_total * 100) if sp_total else 0
            row_data = [carrier, remark, len(remark_group), round(sp_total,2), round(noise_total,2), round(variance,2), f"{variance_pct:.2f}%"]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(current_row, col, val)
                cell.fill = _fill(REMARK_COLORS.get(remark, "FFFFFF"))
                cell.border = _border()
                cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
            current_row += 1

        # Carrier subtotal
        c_sp    = c_group["SP Total"].sum()
        c_noise = c_group["Expected Total"].sum()
        c_var   = c_group["Charges Diff (HKD)"].sum()
        c_pct   = (c_var / c_sp * 100) if c_sp else 0
        subtotal_data = [f"{carrier} TOTAL", "", len(c_group), round(c_sp,2), round(c_noise,2), round(c_var,2), f"{c_pct:.2f}%"]
        for col, val in enumerate(subtotal_data, 1):
            cell = ws.cell(current_row, col, val)
            cell.fill = _fill(COLORS["total_bg"])
            cell.font = _bold_font(size=9)
            cell.border = _border()
        current_row += 1

    # ── Section C: INR conversion note ────────────────────────────────────
    current_row += 1
    ws.cell(current_row, 1, f"Total Variance in HKD:").font = Font(bold=True, size=10)
    ws.cell(current_row, 2, round(grand_var, 2)).font = Font(bold=True, size=10, color="C00000")
    current_row += 1
    inr_rate = 11.65
    ws.cell(current_row, 1, f"Approx in INR (@ ₹{inr_rate}/HKD):").font = Font(bold=True, size=10)
    ws.cell(current_row, 2, f"₹{grand_var * inr_rate:,.0f}").font = Font(bold=True, size=10, color="C00000")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 22

    ws.freeze_panes = "A5"


def _write_detail_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, sheet_title: str, description: str):
    ws = wb.create_sheet(sheet_title)

    # Title and description
    ws["A1"] = sheet_title
    ws["A1"].font = Font(bold=True, size=13, color=COLORS["header_bg"])
    ws["A2"] = description
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    # Headers (row 4)
    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = _fill(COLORS["header_bg"])
        cell.font = _bold_font(COLORS["header_font"], 9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _border()

    # Data rows (start from row 5)
    remarks_col_idx = headers.index("Remarks") + 1 if "Remarks" in headers else None

    for row_idx, (_, row) in enumerate(df.iterrows(), start=5):
        remark = row.get("Remarks", "")
        row_color = REMARK_COLORS.get(remark, "FFFFFF")

        for col_idx, h in enumerate(headers, 1):
            val = row[h]
            # Convert datetime objects to strings
            if hasattr(val, 'strftime'):
                val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row_idx, col_idx, val)
            cell.fill = _fill(row_color)
            cell.border = _border()
            cell.font = Font(size=9)

            # Right-align numbers
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Column widths
    col_widths = {
        "Tracking Number": 26, "Date": 12, "Courier": 32, "Order Ref": 12,
        "Invoice Number": 12, "Invoice Date": 12, "From": 6, "To": 6,
        "Country Name": 18, "SKU": 26, "Qty": 5,
        "SP Final Weight (kg)": 18, "SP Freight": 12, "SP Fuel Surcharge": 18,
        "SP Seasonal": 12, "SP Residential": 14, "SP Remote Area": 14, "SP Total": 12,
        "Product Weight (kg)": 18, "Expected Weight (kg)": 20, "Rate Tier Used (kg)": 18,
        "Expected Freight": 16, "Expected Fuel": 14, "Expected Total": 16,
        "Charges Diff (HKD)": 18, "Weight Diff (kg)": 16, "Freight Diff": 14, "Fuel Diff": 12,
        "Remarks": 14,
    }
    for col_idx, h in enumerate(headers, 1):
        width = col_widths.get(h, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    print(f"  ✓ Sheet '{sheet_title}' written: {len(df):,} rows")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Spaceship Shipping Fee Reconciliation Engine",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Run against the master Excel (re-validation mode):
  python3 reconcile.py --source-excel "Spaceship Shipping Fee Validation - Jan'26.xlsx" --month "Jan 2026"

  # Run against fresh CSV exports (normal monthly mode):
  python3 reconcile.py \\
      --invoice-csv inputs/week1.csv inputs/week2.csv inputs/week3.csv inputs/week4.csv \\
      --order-csv   inputs/order_report.csv \\
      --rates-excel inputs/commercial_rates.xlsx \\
      --product-excel inputs/product_master.xlsx \\
      --month "Feb 2026"
        """
    )

    # Mode A: from master Excel
    parser.add_argument("--source-excel", help="Path to the master Excel file (reads all sheets from it)")

    # Mode B: from CSV exports
    parser.add_argument("--invoice-csv", nargs="+", help="Path(s) to Spaceship invoice CSV(s)")
    parser.add_argument("--order-csv",   help="Path to Noise/Shopify order report CSV")
    parser.add_argument("--rates-excel", help="Path to commercial rates Excel (if not using --source-excel)")
    parser.add_argument("--product-excel", help="Path to product weight master Excel (if not using --source-excel)")

    # Common options
    parser.add_argument("--month",  default="", help='Month label for the report title, e.g. "Jan 2026"')
    parser.add_argument("--output", help="Output Excel file path (default: auto-named)")

    args = parser.parse_args()

    print("=" * 65)
    print("  SPACESHIP SHIPPING FEE RECONCILIATION ENGINE")
    print(f"  Month: {args.month or '(not specified)'}")
    print("=" * 65)

    # ── Determine input paths ──────────────────────────────────────────────
    if args.source_excel:
        rates_path   = args.source_excel
        product_path = args.source_excel
        rates_sheet  = "Commercial"
        product_sheet= "Product LBH Master"
    else:
        if not args.invoice_csv:
            print("ERROR: You must provide either --source-excel or --invoice-csv")
            sys.exit(1)
        # Fall back to the bundled master Excel for rates and product weights
        # if no separate files are specified
        _default_excel = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "Spaceship Shipping Fee Validation - Jan'26.xlsx"
        )
        rates_path    = args.rates_excel   or _default_excel
        product_path  = args.product_excel or _default_excel
        rates_sheet   = "Commercial"
        product_sheet = "Product LBH Master"

    # ── Load rate table ────────────────────────────────────────────────────
    print("\n[1/4] Loading commercial rate table...")
    rates = load_commercial_rates(rates_path, rates_sheet)

    # ── Load product weights ───────────────────────────────────────────────
    print("\n[2/4] Loading product weight master...")
    product_weights = load_product_weights(product_path, product_sheet)

    # ── Load invoice data ──────────────────────────────────────────────────
    print("\n[3/4] Loading invoice data...")
    if args.source_excel:
        invoice_df = load_invoice_from_excel(args.source_excel)
    else:
        invoice_df = load_invoice_from_csv(args.invoice_csv, args.order_csv)

    # ── Run reconciliation ─────────────────────────────────────────────────
    print("\n[4/4] Running reconciliation...")
    result_df = reconcile(invoice_df, rates, product_weights)

    # ── Generate output ────────────────────────────────────────────────────
    if args.output:
        output_path = args.output
    else:
        month_slug  = (args.month or "output").replace(" ", "_").replace("'", "")
        output_path = f"reconciliation_{month_slug}.xlsx"

    generate_report(result_df, output_path, args.month)

    print("\n" + "=" * 65)
    print("  DONE.")
    print(f"  Report: {output_path}")
    print("=" * 65)


if __name__ == "__main__":
    main()
